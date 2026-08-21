"""Descriptive KPI catalog: a read-only, frontend-facing view of the workbook.

This is the *catalogue* projection — what each KPI **is**, how it relates to the
others, its formula, and its literature references. It deliberately carries none
of the compute/company-input machinery (Value, Target Min/Max, Weight, status):
that lives in ``model.Kpi`` / ``model.Result`` and the ``compute`` command.

Output is a flat, normalized ``id -> CatalogKpi`` map plus a ``label -> reference``
bibliography, so a frontend builds whatever tree/view it wants and resolves each
reference code on its own. Keys are camelCased for direct JS consumption.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]
from pydantic import BaseModel, ConfigDict
from pydantic.alias_generators import to_camel

# prefix, sheet, domain-score id
DOMAINS: tuple[tuple[str, str, str], ...] = (
    ("EN", "Environmental Impact", "EN0"),
    ("EC", "Economic Viability", "EC0"),
    ("C", "Circular Efforts", "C0"),
    ("R", "Resource Efficiency", "R0"),
    ("S", "Social Impact", "S0"),
)
_SHEET_PREFIX = {sheet: prefix for prefix, sheet, _ in DOMAINS}
REFERENCES_SHEET = "References"


class _CamelModel(BaseModel):
    """Snake_case attributes in Python, camelCase keys in JSON (for the frontend)."""

    model_config = ConfigDict(alias_generator=to_camel, populate_by_name=True)


class CatalogReference(_CamelModel):
    """One bibliography entry from the References sheet (keyed by its Label)."""

    title: str | None = None
    description: str | None = None
    type: str | None = None
    link: str | None = None


class CatalogKpi(_CamelModel):
    """The descriptive definition of one KPI row."""

    id: str
    domain: str  # EN / EC / C / R / S
    sheet: str
    level: int | None = None
    name: str
    description: str | None = None
    objective: str | None = None
    underlying: list[str] = []  # child ids that feed into this row
    parents: list[str] = []  # ids this row feeds up into
    potential_reference_values: str | None = None
    unit: str | None = None
    formula: str | None = None
    references: list[str] = []  # codes -> resolve in Catalog.references
    calculation_strategy: str | None = None
    is_raw_data_point: bool = False  # the workbook's "Data? = x" flag
    archived: bool = False  # intentionally dormant; not a defect
    data_source: str | None = None
    lifecycle_stages: list[str] = []
    example_value: str | None = None
    comment: str | None = None


class CatalogDomain(_CamelModel):
    id: str
    name: str
    score_id: str


class CatalogMeta(_CamelModel):
    source: str
    note: str = "Descriptive master catalog. No company inputs or computed scores."
    kpi_count: int
    reference_count: int


class Catalog(_CamelModel):
    meta: CatalogMeta
    domains: list[CatalogDomain]
    kpis: dict[str, CatalogKpi]
    references: dict[str, CatalogReference]


def _text(raw: object) -> str | None:
    if raw is None:
        return None
    text = str(raw).strip()
    return text or None


def _split(raw: object) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if text in ("", "None"):
        return []
    return [p.strip() for p in text.replace("\r\n", "\n").split("\n")
            if p.strip() and p.strip() != "None"]


def _as_int(raw: object) -> int | None:
    text = _text(raw)
    if text is None:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def load_catalog(path: str | Path) -> Catalog:
    """Read ``path`` and return the descriptive catalog (master workbook)."""
    wb = load_workbook(filename=str(path), data_only=True)
    kpis: dict[str, CatalogKpi] = {}

    for prefix, sheet, _score in DOMAINS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c).strip()) for c in next(rows)]
        idx = {name: i for i, name in enumerate(header)}

        def cell(row: tuple[object, ...], name: str) -> object:
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        for row in rows:
            kid = _text(cell(row, "#"))
            if not kid:
                continue
            comment = _text(cell(row, "Comment"))
            # header drift: "Example Value" (Environmental) vs "Example Values" (others)
            example = _text(cell(row, "Example Value")) or _text(cell(row, "Example Values"))
            kpis[kid] = CatalogKpi(
                id=kid,
                domain=prefix,
                sheet=sheet,
                level=_as_int(cell(row, "Level")),
                name=_text(cell(row, "Indicator Name")) or kid,
                description=_text(cell(row, "Description")),
                objective=_text(cell(row, "Objective / Goal")),
                underlying=_split(cell(row, "Underlying Metrics")),
                parents=_split(cell(row, "Parent Metrics")),
                potential_reference_values=_text(cell(row, "Potential Reference Values")),
                unit=_text(cell(row, "Unit")),
                formula=_text(cell(row, "Formula")),
                references=_split(cell(row, "Reference")),
                calculation_strategy=_text(cell(row, "Calculation Strategy")),
                is_raw_data_point=str(cell(row, "Data?") or "").strip().lower() == "x",
                archived="archived" in (comment or "").lower(),
                data_source=_text(cell(row, "Data Source")),
                lifecycle_stages=_split(cell(row, "Product Life Cycle Stages")),
                example_value=example,
                comment=comment,
            )

    references = _load_references(wb)
    wb.close()

    return Catalog(
        meta=CatalogMeta(
            source=Path(path).name,
            kpi_count=len(kpis),
            reference_count=len(references),
        ),
        domains=[CatalogDomain(id=p, name=n, score_id=s) for p, n, s in DOMAINS],
        kpis=kpis,
        references=references,
    )


def _load_references(wb: object) -> dict[str, CatalogReference]:
    if REFERENCES_SHEET not in wb.sheetnames:  # type: ignore[attr-defined]
        return {}
    ws = wb[REFERENCES_SHEET]  # type: ignore[index]
    rows = ws.iter_rows(values_only=True)
    header = [("" if c is None else str(c).strip()) for c in next(rows)]
    idx = {name: i for i, name in enumerate(header)}

    def cell(row: tuple[object, ...], name: str) -> object:
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    out: dict[str, CatalogReference] = {}
    for row in rows:
        label = _text(cell(row, "Label"))
        if not label:
            continue
        out[label] = CatalogReference(
            title=_text(cell(row, "Title")),
            description=_text(cell(row, "Description")),
            type=_text(cell(row, "Type")),
            link=_text(cell(row, "Link")),
        )
    return out
