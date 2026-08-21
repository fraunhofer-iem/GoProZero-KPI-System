"""Load the KPI workbook into the typed model.

Reads only the five domain sheets (the source of truth). Columns are located by header
name, not position, so the loader survives column reordering. Builds a GLOBAL id->Kpi map
across all domains so cross-domain children (e.g. EC5 -> EN1-4, R2-7) resolve.
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from kpi_engine.model import Kpi, Strategy

DOMAIN_SHEETS: tuple[str, ...] = (
    "Environmental Impact", "Economic Viability", "Circular Efforts",
    "Resource Efficiency", "Social Impact",
)

# header label -> attribute. Inputs (Value/Reference Value/Weight/Target *) are optional.
_TEXT_COLS = {"Indicator Name": "name", "Unit": "unit", "Formula": "formula_text"}
_NUM_COLS = {"Weight": "weight", "Target Min": "target_min", "Target Max": "target_max",
             "Value": "value", "Reference Value": "reference_value"}


def _split(raw: object) -> list[str]:
    if raw is None:
        return []
    text = str(raw).strip()
    if text in ("", "None"):
        return []
    return [part.strip() for part in text.replace("\r\n", "\n").split("\n") if part.strip()]


def _as_float(raw: object) -> float | None:
    if raw is None or (isinstance(raw, str) and raw.strip() == ""):
        return None
    try:
        return float(raw)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return None


def _as_int(raw: object) -> int | None:
    f = _as_float(raw)
    return int(f) if f is not None else None


def _as_strategy(raw: object) -> Strategy | None:
    text = "" if raw is None else str(raw).strip()
    try:
        return Strategy(text)
    except ValueError:
        return None


def load_workbook_model(path: str | Path) -> dict[str, Kpi]:
    """Read ``path`` and return a global ``{kpi_id: Kpi}`` map across all domain sheets."""
    # Not read_only: the surgical editor (tools/scripts/xlsx_edit.py) appends cells in the
    # right-hand input block (Target Min/Max, Value, Reference Value) without rewriting the
    # worksheet <dimension>, and read_only mode trusts that stale dimension and silently drops
    # the appended columns. Normal mode reads every cell present. Negligible cost at this size.
    wb = load_workbook(filename=str(path), data_only=True)
    kpis: dict[str, Kpi] = {}
    for sheet in DOMAIN_SHEETS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        rows = ws.iter_rows(values_only=True)
        header = [("" if c is None else str(c).strip()) for c in next(rows)]
        idx = {name: i for i, name in enumerate(header)}

        def cell(row: tuple[object, ...], name: str) -> object:
            i = idx.get(name)
            return row[i] if i is not None and i < len(row) else None

        for row in rows:
            kid = cell(row, "#")
            kid = "" if kid is None else str(kid).strip()
            if not kid:
                continue
            data_flag = str(cell(row, "Data?") or "").strip().lower() == "x"
            archived = "archived" in str(cell(row, "Comment") or "").lower()
            kpi = Kpi(
                id=kid,
                name=str(cell(row, "Indicator Name") or "").strip(),
                sheet=sheet,
                strategy=_as_strategy(cell(row, "Calculation Strategy")),
                children=_split(cell(row, "Underlying Metrics")),
                parents=_split(cell(row, "Parent Metrics")),
                level=_as_int(cell(row, "Level")),
                is_data=data_flag,
                archived=archived,
                unit=(str(cell(row, "Unit")).strip() or None) if cell(row, "Unit") else None,
                formula_text=(str(cell(row, "Formula")).strip() or None) if cell(row, "Formula") else None,
                weight=_as_float(cell(row, "Weight")),
                target_min=_as_float(cell(row, "Target Min")),
                target_max=_as_float(cell(row, "Target Max")),
                value=_as_float(cell(row, "Value")),
                reference_value=_as_float(cell(row, "Reference Value")),
            )
            kpis[kid] = kpi
    wb.close()
    return kpis
