#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Print all sheets of an xlsx to stdout as text. Used as a git textconv diff driver
so that `git diff` on the workbook is readable on the command line."""
import sys
from openpyxl import load_workbook

# On Windows, stdout defaults to cp1252 and chokes on characters like "₂".
# Git captures this output for the diff, so force UTF-8.
sys.stdout.reconfigure(encoding="utf-8")


def cell_text(v):
    if v is None:
        return ""
    return (
        str(v)
        .replace("\t", "\\t")
        .replace("\r\n", "\\n")
        .replace("\n", "\\n")
        .replace("\r", "\\n")
    )


def main(path):
    wb = load_workbook(path, data_only=False)
    for sheet in wb.sheetnames:
        print(f"## {sheet}")
        for row in wb[sheet].iter_rows():
            print("\t".join(cell_text(c.value) for c in row).rstrip("\t"))
        print()


if __name__ == "__main__":
    main(sys.argv[1])
