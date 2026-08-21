#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1", "lxml>=5"]
# ///
"""Seed equal-weight defaults into the master workbook's 'Weight' column.

Policy (T1.1 of reviews/KPI-system-gaps.md)
-------------------------------------------
The workbook ships with an empty Weight column, so no WEIGHTED_AVERAGE /
WEIGHTED_RATIO score can compute. This writes a neutral default so the system
produces a score the moment it is opened, while a company overrides any cell.

- A weight is written ONLY on rows that are a child of a WEIGHTED_AVERAGE_STRATEGY
  or WEIGHTED_RATIO_STRATEGY parent. Children of NORMALIZED_RATIO parents feed a
  ratio, not a weighted average, so they get NO weight (left blank).
- Default value = 1/N, where N = the number of *existing* listed children of the
  row's PRIMARY parent. The last sibling absorbs rounding so each parent's
  children sum to exactly 1.0.
- A row with two weighted parents has only one Weight cell. Its primary parent is
  the one sharing the longest ID prefix (structural nesting); ties -> shorter
  parent id, then lexical. Any weighted parent left unable to sum to 1 under this
  rule is REPORTED (these are the Tier-2 structural anomalies).

Writes via the surgical editor so threaded comments / colours / hyperlinks are
preserved. Output goes to output/ ; promote to data/ manually after review.

Usage:
    uv run tools/scripts/seed_default_weights.py                     # report only
    uv run tools/scripts/seed_default_weights.py --write             # write to output/
"""
from __future__ import annotations
import argparse
import os
import sys
from collections import defaultdict

from openpyxl import load_workbook

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from xlsx_edit import set_cells  # noqa: E402

DOMAINS = ["Environmental Impact", "Economic Viability", "Circular Efforts",
           "Resource Efficiency", "Social Impact"]
WSTRAT = {"WEIGHTED_AVERAGE_STRATEGY", "WEIGHTED_RATIO_STRATEGY"}
ID_COL, UND_COL, STRAT_COL, WEIGHT_COL = 1, 5, 15, 17  # 1-based; Weight = col Q(17)
WEIGHT_REF = "Q"


def multi(v):
    v = ("" if v is None else str(v)).strip()
    if v in ("", "None"):
        return []
    return [x.strip() for x in v.replace("\r\n", "\n").split("\n") if x.strip()]


def common_prefix_len(a, b):
    n = 0
    for x, y in zip(a, b):
        if x != y:
            break
        n += 1
    return n


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", default="data/KPI List.xlsx")
    ap.add_argument("--out", default="output/KPI List.weights.xlsx")
    ap.add_argument("--write", action="store_true", help="write the edited workbook to --out")
    args = ap.parse_args()

    wb = load_workbook(args.src, data_only=False)

    # id -> (sheet, row, strategy, [children]); also row lookup per sheet
    info = {}
    for d in DOMAINS:
        ws = wb[d]
        for r in range(2, ws.max_row + 1):
            rid = ws.cell(r, ID_COL).value
            rid = "" if rid is None else str(rid).strip()
            if not rid:
                continue
            strat = (ws.cell(r, STRAT_COL).value or "")
            und = ws.cell(r, UND_COL).value
            info[rid] = {"sheet": d, "row": r, "strat": str(strat).strip(),
                         "children": multi(und)}

    ids = set(info)
    # weighted parent -> existing children (in listed order)
    wparents = {}
    for pid, rec in info.items():
        if rec["strat"] in WSTRAT:
            wparents[pid] = [c for c in rec["children"] if c in ids]

    # child -> list of weighted parents
    child_wp = defaultdict(list)
    for pid, kids in wparents.items():
        for c in kids:
            child_wp[c].append(pid)

    # resolve primary parent per child
    primary = {}
    for c, ps in child_wp.items():
        ps_sorted = sorted(ps, key=lambda p: (-common_prefix_len(c, p), len(p), p))
        primary[c] = ps_sorted[0]

    # weight written on each child = 1/N of its primary parent (last sibling absorbs rounding)
    weight_of = {}
    for pid, kids in wparents.items():
        owned = [c for c in kids if primary.get(c) == pid]
        # we still need a value for every listed child; base on primary parent's N
        # (compute per primary parent below)
    # compute by primary parent so each parent's owned children sum to 1
    by_primary = defaultdict(list)
    for c, p in primary.items():
        by_primary[p].append(c)
    for p, owned in by_primary.items():
        # order them as they appear under p
        ordered = [c for c in wparents[p] if c in owned]
        n = len(ordered)
        if n == 0:
            continue
        base = round(1.0 / n, 4)
        for i, c in enumerate(ordered):
            weight_of[c] = base if i < n - 1 else round(1.0 - base * (n - 1), 4)

    # ---- report: does each weighted parent's listed children sum to 1? ----
    print(f"Weighted parents: {len(wparents)} | children receiving a default weight: {len(weight_of)}\n")
    bad = []
    for pid, kids in sorted(wparents.items()):
        listed = info[pid]["children"]
        missing = [c for c in listed if c not in ids]
        s = round(sum(weight_of.get(c, 0.0) for c in kids), 4)
        if abs(s - 1.0) > 0.001 or missing:
            bad.append((pid, s, kids, missing))
    if bad:
        print("PARENTS whose default child-weights do NOT sum to 1 (need manual attention):")
        for pid, s, kids, missing in bad:
            extra = f" | lists missing child(ren) {missing}" if missing else ""
            shared = [c for c in kids if primary.get(c) != pid]
            note = f" | {len(shared)} child(ren) primary-assigned to another parent" if shared else ""
            print(f"   {pid} ({info[pid]['strat'].split('_')[0]}): sum={s}{extra}{note}")
        print()
    else:
        print("All weighted parents sum to 1.0 under the default.\n")

    # ---- build edits ----
    edits = defaultdict(dict)
    for c, w in weight_of.items():
        rec = info[c]
        edits[rec["sheet"]][f"{WEIGHT_REF}{rec['row']}"] = w
    total = sum(len(v) for v in edits.values())
    print(f"Default weights to write: {total} cells across {len(edits)} sheets")
    for sh in DOMAINS:
        if sh in edits:
            print(f"   {sh}: {len(edits[sh])} cells")

    if args.write:
        os.makedirs(os.path.dirname(args.out), exist_ok=True)
        set_cells(args.src, args.out, {k: v for k, v in edits.items()})
        print(f"\nWrote {args.out} (comments/colours/hyperlinks preserved). "
              f"Review, then promote to data/ manually.")
    else:
        print("\n(report only — pass --write to produce the edited workbook in output/)")


if __name__ == "__main__":
    main()
