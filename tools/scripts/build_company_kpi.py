#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Build a company-specific KPI workbook by pruning out-of-scope KPIs.

The master workbook (data/KPI List.xlsx) is the full KPI tree. A workshop decides
which KPIs are out of scope for a given product/company; that decision is recorded
in a **scope file** and this script removes those KPIs *and their now-orphaned
descendants*, writing a derived workbook to output/.

Scope files (kept local, not versioned)
---------------------------------------
The per-company scope lives outside the code, in ``--scope-dir`` (default
``data/others/``, which is git-ignored). Each is ``<company>.scope.json``:

    {
      "label":  "Acme",                       # used for the output filename
      "remove": { "EN4": "why this is out of scope", ... },   # OUT decisions
      "review": { "EN5": "ambiguous — left in, flagged", ... },# reported, kept
      "annotate": { "C34": "workshop note -> Comment cell", ... }  # optional
    }

Only ``label`` and ``remove`` are required; ``review`` and ``annotate`` may be
omitted. Run with ``uv run tools/scripts/build_company_kpi.py acme`` to load
``data/others/acme.scope.json``, or point ``--scope`` at any file directly.

How removal cascades
--------------------
A KPI is KEPT iff it is still reachable from the domain root (EN0/EC0/C0/R0/S0)
through nodes that are NOT in the removed set. Child links are taken from BOTH the
'Underlying Metrics' (col 5) and the inverse of 'Parent Metrics' (col 6), so a raw
metric shared by a kept KPI survives even if one of its parents is removed.

For every kept aggregate, its 'Underlying Metrics' cell (col 5) is rewritten to list
only the children that survived, and its 'Parent Metrics' cell (col 6) is rewritten
to drop any parent that was removed in *any* domain — so no cell points at a KPI that
is no longer in the workbook. (The Metrics List's Parent Metrics column is a
=HYPERLINK() navigation formula, not a plain id list, so it is left untouched.)

What is preserved / lost
------------------------
The derived file is the master, edited in place and saved to output/ — NOT a fresh
rebuild. openpyxl's load->save preserves cell values, fill colours (incl. the
theme-indexed ones, so no palette drift), fonts, borders, alignment, number formats,
merged cells, row heights, column widths and hyperlinks. The static sheets (Overview,
Top-Level, References) are copied verbatim; only the domain sheets + Metrics List have
out-of-scope rows deleted.
NOTE: openpyxl cannot round-trip two things, which ``restore_images`` / the model
handle explicitly: anchored **images** (e.g. the Top-Level diagram) are dropped on
save and grafted back afterwards; the master's *threaded comments* are dropped and
NOT carried over (the textual 'Comment' column, col 18, is preserved). The canonical
data/ workbook is opened read-only and never touched.

Usage:
    uv run tools/scripts/build_company_kpi.py acme
    uv run tools/scripts/build_company_kpi.py acme --src "data/KPI List.xlsx" --out output/
    uv run tools/scripts/build_company_kpi.py --scope path/to/acme.scope.json
"""
from __future__ import annotations

import argparse
import json
import os
import posixpath
import re
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

from openpyxl import load_workbook

DOMAINS = ["Environmental Impact", "Economic Viability", "Circular Efforts",
           "Resource Efficiency", "Social Impact"]
DOMAIN_ROOTS = {"Environmental Impact": "EN0", "Economic Viability": "EC0",
                "Circular Efforts": "C0", "Resource Efficiency": "R0",
                "Social Impact": "S0"}
COPY_AS_IS = ["Overview", "Top-Level", "References"]  # not pruned
SHEET_ORDER = ["Overview", "Top-Level", "Environmental Impact", "Economic Viability",
               "Circular Efforts", "Resource Efficiency", "Social Impact",
               "Metrics List", "References"]

ID_COL, CHILD_COL, PARENT_COL, LEVEL_COL, DATA_COL, COMMENT_COL = 1, 5, 6, 11, 12, 18

DEFAULT_SCOPE_DIR = "data/others"


def load_scope(company: str | None, scope: str | None, scope_dir: str) -> dict:
    """Load and validate a per-company scope file.

    Either ``scope`` (an explicit path) or ``company`` (resolved to
    ``<scope_dir>/<company>.scope.json``) must be given.
    """
    if scope:
        path = Path(scope)
    elif company:
        path = Path(scope_dir) / f"{company}.scope.json"
    else:  # argparse guards this, but be explicit
        raise SystemExit("error: provide a company name or --scope <file>")

    if not path.exists():
        raise SystemExit(
            f"error: scope file not found: {path}\n"
            f"Create it (see the module docstring for the schema) or pass --scope.")

    cfg = json.loads(path.read_text(encoding="utf-8"))
    if "label" not in cfg or "remove" not in cfg:
        raise SystemExit(f"error: {path} must contain at least 'label' and 'remove'.")
    cfg.setdefault("review", {})
    cfg.setdefault("annotate", {})
    return cfg


def _split(v):
    """Split a multi-value cell (newline-separated) into clean tokens."""
    if v in (None, "", "None"):
        return []
    return [t.strip() for t in str(v).split("\n") if t.strip() and t.strip() != "None"]


def _row_id(row):
    v = row[ID_COL - 1].value
    if v in (None, "") or str(v).startswith("#"):
        return None
    return str(v).strip()


def compute_kept(ws, root):
    """Return (kept_ids, child_map, id_to_row) for one domain sheet."""
    rows = {}            # id -> openpyxl row tuple
    child_map = {}       # id -> ordered list of child ids (col5 ∪ inverse col6)
    for row in ws.iter_rows(min_row=2):
        rid = _row_id(row)
        if not rid:
            continue
        rows[rid] = row
        child_map.setdefault(rid, [])
        for c in _split(row[CHILD_COL - 1].value):
            if c not in child_map[rid]:
                child_map[rid].append(c)
    # union with inverse of Parent Metrics
    for rid, row in rows.items():
        for p in _split(row[PARENT_COL - 1].value):
            child_map.setdefault(p, [])
            if rid not in child_map[p]:
                child_map[p].append(rid)
    return rows, child_map


def reachable(root, child_map, removed):
    """BFS from root through child_map, skipping removed nodes."""
    if root in removed:
        return set()
    kept, stack = {root}, [root]
    while stack:
        node = stack.pop()
        for c in child_map.get(node, []):
            if c in removed or c in kept:
                continue
            kept.add(c)
            stack.append(c)
    return kept


def _norm_id(v):
    """Normalize a col-A cell value to a clean id, or None for blank/comment rows."""
    if v in (None, "") or str(v).strip().startswith("#"):
        return None
    return str(v).strip()


def prune_rows(ws, keep_ids):
    """Delete every data row (row >= 2) whose col-A id is not in keep_ids; keep the header."""
    doomed = [r for r in range(2, ws.max_row + 1)
              if _norm_id(ws.cell(r, ID_COL).value) not in keep_ids]
    for r in reversed(doomed):  # bottom-up so earlier indices stay valid
        ws.delete_rows(r, 1)


# ---------------------------------------------------------------------------
# Post-save image restore
# ---------------------------------------------------------------------------
# openpyxl's load->save round-trip preserves the theme, styles, merges, row heights
# and the untouched static sheets — but it cannot round-trip anchored IMAGES
# (DrawingML): the Top-Level diagram lives in xl/drawings/ + xl/media/ and is
# dropped on save. We reopen the saved .xlsx (a zip) and graft those parts back,
# wiring them only onto COPY_AS_IS sheets — their rows aren't pruned, so the cell
# anchors stay valid. (Threaded comments are also dropped, as before — see docs.)
_RELS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
_DRAWING_CT = "application/vnd.openxmlformats-officedocument.drawing+xml"


def _rels_path(part):
    """Path of the .rels part for an OOXML part (xl/a/b.xml -> xl/a/_rels/b.xml.rels)."""
    return f"{posixpath.dirname(part)}/_rels/{posixpath.basename(part)}.rels"


def _sheet_parts(zf):
    """Map worksheet display-name -> part path (e.g. 'Top-Level' -> xl/worksheets/sheet2.xml)."""
    wb = ET.fromstring(zf.read("xl/workbook.xml"))
    rid_to_name = {s.get(f"{{{_R_NS}}}id"): s.get("name")
                   for s in wb.iter(f"{{{_MAIN_NS}}}sheet")}
    rels = ET.fromstring(zf.read("xl/_rels/workbook.xml.rels"))
    out = {}
    for rel in rels.findall(f"{{{_RELS_NS}}}Relationship"):
        if rel.get("Type", "").endswith("/worksheet"):
            tgt = rel.get("Target").lstrip("/")
            if not tgt.startswith("xl/"):
                tgt = "xl/" + tgt
            out[rid_to_name.get(rel.get("Id"))] = tgt
    return out


def _collect_drawings(mz, mnames, sheet_part):
    """For one master worksheet part, collect its drawing part(s) + rels + media as {path: bytes}."""
    relp = _rels_path(sheet_part)
    if relp not in mnames:
        return []
    grafts = []
    for rel in ET.fromstring(mz.read(relp)).findall(f"{{{_RELS_NS}}}Relationship"):
        if not rel.get("Type", "").endswith("/drawing"):
            continue
        dpart = posixpath.normpath(posixpath.join(posixpath.dirname(sheet_part),
                                                  rel.get("Target")))
        if dpart not in mnames:
            continue
        parts = {dpart: mz.read(dpart)}
        drelp = _rels_path(dpart)
        if drelp in mnames:
            parts[drelp] = mz.read(drelp)
            for r2 in ET.fromstring(parts[drelp]).findall(f"{{{_RELS_NS}}}Relationship"):
                media = posixpath.normpath(posixpath.join(posixpath.dirname(dpart),
                                                          r2.get("Target")))
                if media in mnames:
                    parts[media] = mz.read(media)
        grafts.append((dpart, parts))
    return grafts


def _add_rel(entries, relp, rtype, target):
    """Append a relationship to a .rels part (creating it if absent); return the new rId."""
    if relp in entries:
        xml = entries[relp].decode("utf-8")
        ids = [int(m) for m in re.findall(r'Id="rId(\d+)"', xml)]
        rid = f"rId{max(ids) + 1 if ids else 1}"
        xml = xml.replace("</Relationships>",
                          f'<Relationship Id="{rid}" Type="{rtype}" Target="{target}"/></Relationships>')
    else:
        rid = "rId1"
        xml = ('<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
               f'<Relationships xmlns="{_RELS_NS}">'
               f'<Relationship Id="{rid}" Type="{rtype}" Target="{target}"/></Relationships>')
    entries[relp] = xml.encode("utf-8")
    return rid


def _add_drawing_tag(sheet_bytes, rid):
    """Declare the r: namespace (if needed) and append <drawing r:id=.../> before </worksheet>."""
    xml = sheet_bytes.decode("utf-8")
    if "xmlns:r=" not in xml.split(">", 1)[0]:
        xml = xml.replace("<worksheet ", f'<worksheet xmlns:r="{_R_NS}" ', 1)
    return xml.replace("</worksheet>", f'<drawing r:id="{rid}"/></worksheet>').encode("utf-8")


def _patch_content_types(ct_bytes, drawing_parts, need_png):
    xml = ct_bytes.decode("utf-8")
    add = ""
    if need_png and 'Extension="png"' not in xml:
        add += '<Default Extension="png" ContentType="image/png"/>'
    for part in drawing_parts:
        if f'PartName="/{part}"' not in xml:
            add += f'<Override PartName="/{part}" ContentType="{_DRAWING_CT}"/>'
    return (xml.replace("</Types>", add + "</Types>") if add else xml).encode("utf-8")


def restore_images(master_path, out_path, copy_as_is):
    """Graft copy-as-is-sheet images back into a workbook openpyxl saved without them."""
    with zipfile.ZipFile(master_path) as mz:
        mnames = set(mz.namelist())
        msheets = _sheet_parts(mz)
        grafts = []  # (sheet_name, drawing_part, {path: bytes})
        for name in copy_as_is:
            spart = msheets.get(name)
            if spart:
                for dpart, parts in _collect_drawings(mz, mnames, spart):
                    grafts.append((name, dpart, parts))

    if not grafts:
        return

    with zipfile.ZipFile(out_path) as oz:
        osheets = _sheet_parts(oz)
        entries = {n: oz.read(n) for n in oz.namelist()}

    drawing_parts, need_png = [], False
    for name, dpart, parts in grafts:
        out_sheet = osheets.get(name)
        if not out_sheet or out_sheet not in entries:
            continue
        for path, data in parts.items():  # graft parts under their master names (no collisions)
            entries[path] = data
            need_png = need_png or path.lower().endswith(".png")
        drawing_parts.append(dpart)
        rid = _add_rel(entries, _rels_path(out_sheet),
                       f"{_R_NS}/drawing", "../drawings/" + posixpath.basename(dpart))
        entries[out_sheet] = _add_drawing_tag(entries[out_sheet], rid)

    entries["[Content_Types].xml"] = _patch_content_types(
        entries["[Content_Types].xml"], drawing_parts, need_png)

    tmp = f"{out_path}.tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zf:
        # [Content_Types].xml conventionally leads the package
        for name in sorted(entries, key=lambda n: (n != "[Content_Types].xml", n)):
            zf.writestr(name, entries[name])
    os.replace(tmp, out_path)


def build(cfg, src_path, out_dir):
    removed_seed = cfg["remove"]
    annotate = cfg.get("annotate") or {}
    # Load the master and edit it in place: openpyxl's load->save preserves the theme,
    # styles, merges, row heights and — crucially — the static sheets (Overview,
    # Top-Level, References) exactly as authored. We only prune rows from the domain
    # sheets + Metrics List, so the sheets that don't change are copied verbatim.
    wb = load_workbook(src_path)

    kept_raw_ids = set()       # raw (Data?=x) ids kept across all domains
    summary = []               # (sheet, kept_count, removed_ids)

    # Phase 1: resolve kept/removed for every domain up front, snapshotting the cell
    # values we'll rewrite (children/parents/data flag) BEFORE any deletion — deleting
    # rows detaches the tuples compute_kept returned. We need the GLOBAL removed set
    # now too, because a kept row's Parent Metrics can name a parent that lives in —
    # and was pruned from — another domain (e.g. a Resource metric under an Env KPI).
    solve = {}                 # sheet -> (child_map, kept, removed_closure, info)
    removed_all = set()        # every id pruned, across all domains
    for sheet in DOMAINS:
        if sheet not in wb.sheetnames:
            continue
        rows, child_map = compute_kept(wb[sheet], DOMAIN_ROOTS[sheet])
        all_ids = set(rows)
        removed = {rid for rid in removed_seed if rid in all_ids}
        kept = reachable(DOMAIN_ROOTS[sheet], child_map, removed)
        info = {rid: (_split(row[CHILD_COL - 1].value),
                      _split(row[PARENT_COL - 1].value),
                      row[DATA_COL - 1].value) for rid, row in rows.items()}
        solve[sheet] = (child_map, kept, all_ids - kept, info)
        removed_all |= (all_ids - kept)
        for rid in kept:  # kept raw (Data?=x) metrics feed Metrics List
            flag = info.get(rid, (None, None, None))[2]
            if flag is not None and str(flag).strip().lower() == "x":
                kept_raw_ids.add(rid)

    # Phase 2: prune each domain sheet in place, then rewrite the surviving
    # aggregates' Underlying Metrics (col 5, surviving children — keeping original
    # order, then appending kept children the master omitted from col 5, e.g. EC4
    # omits EC46/EC47) and Parent Metrics (col 6, minus any id removed in ANY domain,
    # leaving genuine cross-domain / external refs), and append workshop annotations.
    for sheet in DOMAINS:
        if sheet not in wb.sheetnames:
            continue
        ws = wb[sheet]
        child_map, kept, removed_closure, info = solve[sheet]
        prune_rows(ws, kept)
        child_override, parent_override = {}, {}
        for rid, (orig_children, orig_parents, _flag) in info.items():
            if rid not in kept:
                continue
            # Filter against BOTH the domain's kept set AND the global removed set:
            # a cross-domain child (e.g. EC5 -> R2-7/EN1-4) is pulled into this domain's
            # `kept` via reachability from its own child list, yet may have been pruned
            # from the domain it actually lives in. removed_all catches those so no
            # surviving cell points at a row deleted elsewhere (mirrors the parent rewrite).
            children = [c for c in orig_children if c in kept and c not in removed_all]
            for c in child_map.get(rid, []):
                if (c in kept and c not in removed_all and c in info
                        and c not in orig_children and c not in children):
                    children.append(c)
            child_override[rid] = children
            pruned_parents = [p for p in orig_parents if p not in removed_all]
            if pruned_parents != orig_parents:
                parent_override[rid] = pruned_parents
        for r in range(2, ws.max_row + 1):
            rid = _norm_id(ws.cell(r, ID_COL).value)
            if rid in child_override and _split(ws.cell(r, CHILD_COL).value):
                kids = child_override[rid]
                ws.cell(r, CHILD_COL).value = "\n".join(kids) if kids else "None"
            if rid in parent_override and _split(ws.cell(r, PARENT_COL).value):
                ps = parent_override[rid]
                ws.cell(r, PARENT_COL).value = "\n".join(ps) if ps else "None"
            if rid in annotate:
                cell = ws.cell(r, COMMENT_COL)
                existing = "" if cell.value in (None, "", "None") else str(cell.value)
                cell.value = f"{existing}\n{annotate[rid]}".strip()
        summary.append((sheet, len(kept) - 1, sorted(removed_closure)))

    # Metrics List: flat mirror of the raw metrics — keep only surviving raw ids. Its
    # Parent Metrics column is a =HYPERLINK() navigation formula, left as-is.
    if "Metrics List" in wb.sheetnames:
        ws = wb["Metrics List"]
        prune_rows(ws, kept_raw_ids)
        summary.append(("Metrics List", ws.max_row - 1, []))

    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, f"{cfg['label']} KPI List.xlsx")
    wb.save(out_path)
    # load->save drops anchored images (openpyxl can't round-trip them); graft the
    # copy-as-is-sheet images back. Threaded comments are dropped too — see docs.
    restore_images(src_path, out_path, COPY_AS_IS)
    return out_path, summary, cfg


def main():
    ap = argparse.ArgumentParser(description="Build a company-specific KPI workbook.")
    ap.add_argument("company", nargs="?",
                    help="scope slug -> <scope-dir>/<company>.scope.json")
    ap.add_argument("--scope", help="explicit path to a .scope.json file")
    ap.add_argument("--scope-dir", default=DEFAULT_SCOPE_DIR,
                    help=f"directory of scope files (default: {DEFAULT_SCOPE_DIR})")
    ap.add_argument("--src", default="data/KPI List.xlsx")
    ap.add_argument("--out", default="output")
    args = ap.parse_args()

    if not args.company and not args.scope:
        ap.error("provide a company name or --scope <file>")

    cfg = load_scope(args.company, args.scope, args.scope_dir)
    out_path, summary, cfg = build(cfg, args.src, args.out)

    print(f"Built {cfg['label']} KPI system -> {out_path}\n")
    print("Pruned per sheet (KPIs kept | removed incl. cascaded descendants):")
    total_removed = 0
    for sheet, kept_n, removed in summary:
        if removed:
            total_removed += len(removed)
            print(f"  {sheet:<22} kept {kept_n:>3}  removed {len(removed):>2}: {', '.join(removed)}")
        else:
            print(f"  {sheet:<22} kept {kept_n:>3}")
    print(f"\nSeed OUT decisions: {len(cfg['remove'])}  (cascaded to {total_removed} rows across domains)")

    annotate = cfg.get("annotate") or {}
    if annotate:
        print(f"\n[ANNOTATED] workshop priority/context written to the Comment cell of {len(annotate)} KPI(s):")
        for rid, note in annotate.items():
            print(f"  {rid:<5} {note}")

    review = cfg.get("review") or {}
    if review:
        print("\n[REVIEW] left IN the workbook but flagged — decide keep/drop:")
        for rid, reason in review.items():
            print(f"  {rid:<5} {reason}")


if __name__ == "__main__":
    main()
