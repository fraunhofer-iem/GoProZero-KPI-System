#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1", "lxml>=5"]
# ///
"""Regenerate the Metrics List sheet from the domain sheets (the source of truth).

Data model: each raw metric (a domain row with Data? = 'x') is the authoritative
record. Metrics List is a flat MIRROR of those raw metrics and must not be hand-edited.
This tool copies every authoritative data field from the domain row into the matching
Metrics List row, so the two can never drift.

What it syncs: all data columns EXCEPT the navigation column (col 6, 'Parent Metrics' on
every sheet — on Metrics List it holds a hyperlink that is left untouched) and the ID column.
'Data Source' (col 13) carries the same name on both (formerly 'Source' on Metrics List).

It edits only cells that actually differ, via the surgical editor (so threaded comments,
colours and the existing hyperlinks are preserved). Rows are matched by ID; metrics that
exist as domain leaves but are missing from Metrics List are reported, not invented.

Usage:
    uv run tools/scripts/sync_metrics_list.py SRC.xlsx DST.xlsx          # write synced copy
    uv run tools/scripts/sync_metrics_list.py SRC.xlsx --report          # dry run, list changes
"""
from __future__ import annotations
import argparse
import os
import sys

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(__file__))
from xlsx_edit import set_cells  # noqa: E402

DOMAINS = ["Environmental Impact", "Economic Viability", "Circular Efforts",
           "Resource Efficiency", "Social Impact"]  # Top-Level has no Data?=x leaves
# Columns to mirror (1-based). Skip 1 (ID, identical) and 6 (Superior/Parent nav hyperlink).
SYNC_COLS = [2, 3, 4, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]
DATA_COL = 12  # 'Data?'


def _norm(v):
    return "" if v is None else str(v)


def compute_sync(path: str):
    """Return (edits, changes, missing, orphans) for the Metrics List sheet."""
    wb = load_workbook(path)

    leaves = {}  # id -> {col: value}
    for d in DOMAINS:
        ws = wb[d]
        for row in ws.iter_rows(min_row=2):
            idv = row[0].value
            if idv in (None, "") or str(idv).startswith("#"):
                continue
            flag = row[DATA_COL - 1].value
            if flag is None or str(flag).strip().lower() != "x":
                continue
            leaves[str(idv).strip()] = {c: row[c - 1].value for c in SYNC_COLS}

    ml = wb["Metrics List"]
    ml_rows = {}
    for row in ml.iter_rows(min_row=2):
        idv = row[0].value
        if idv not in (None, "") and not str(idv).startswith("#"):
            ml_rows[str(idv).strip()] = row[0].row

    edits = {}
    changes = []  # (id, column-name, old, new)
    for idv, vals in leaves.items():
        if idv not in ml_rows:
            continue
        r = ml_rows[idv]
        for c, newv in vals.items():
            cur = ml.cell(row=r, column=c).value
            if _norm(cur) != _norm(newv):
                ref = f"{get_column_letter(c)}{r}"
                edits[ref] = "" if newv is None else newv
                changes.append((idv, ml.cell(row=1, column=c).value, _norm(cur), _norm(newv)))

    missing = sorted(set(leaves) - set(ml_rows))      # domain leaf, no Metrics List row
    orphans = sorted(set(ml_rows) - set(leaves))      # Metrics List row, not a domain leaf
    return {"Metrics List": edits}, changes, missing, orphans


def main():
    ap = argparse.ArgumentParser(description="Regenerate Metrics List from the domain sheets.")
    ap.add_argument("src")
    ap.add_argument("dst", nargs="?")
    ap.add_argument("--report", action="store_true", help="dry run; print changes, don't write")
    args = ap.parse_args()

    edits, changes, missing, orphans = compute_sync(args.src)
    n = len(edits["Metrics List"])
    print(f"Metrics List sync: {n} cell(s) differ from the domain source of truth.")
    by_col = {}
    for _id, col, _o, _new in changes:
        by_col[col] = by_col.get(col, 0) + 1
    for col, cnt in sorted(by_col.items(), key=lambda x: -x[1]):
        print(f"  {cnt:>3}  {col}")
    for _id, col, o, nw in changes[:25]:
        print(f"    {_id} | {col}: {o!r} -> {nw!r}")
    if len(changes) > 25:
        print(f"    ... and {len(changes) - 25} more")
    if missing:
        print(f"\n[!] {len(missing)} domain leaf metric(s) MISSING from Metrics List "
              f"(not added automatically): {', '.join(missing)}")
    if orphans:
        print(f"\n[!] {len(orphans)} Metrics List row(s) with no domain leaf: {', '.join(orphans)}")

    if args.report or not args.dst:
        print("\n(report only — no file written)")
        return
    set_cells(args.src, args.dst, edits)
    print(f"\nWrote synced workbook to {args.dst}")


if __name__ == "__main__":
    main()
