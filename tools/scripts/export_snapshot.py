#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.10"
# dependencies = ["openpyxl>=3.1"]
# ///
"""Export each worksheet of the KPI workbook to a stable TSV snapshot for git diffing.

One spreadsheet row == one text line, so git produces clean line-level diffs.
Captures cell values and formulas (e.g. =HYPERLINK targets). Does not capture
cell fill colour; hierarchy level is tracked via the workbook's own 'Level' column.

Run with: uv run tools/scripts/export_snapshot.py
"""
import os
from openpyxl import load_workbook

WORKBOOK = os.environ.get("KPI_WORKBOOK", "data/KPI List.xlsx")
OUT_DIR = "snapshot"


def cell_text(v):
    if v is None:
        return ""
    s = str(v)
    # Keep every spreadsheet row on a single line and keep whitespace stable.
    return (
        s.replace("\\", "\\\\")
        .replace("\t", "\\t")
        .replace("\r\n", "\\n")
        .replace("\n", "\\n")
        .replace("\r", "\\n")
    )


def safe_name(name):
    return "".join(c if (c.isalnum() or c in " -_") else "_" for c in name).strip()


def main():
    wb = load_workbook(WORKBOOK, data_only=False)
    os.makedirs(OUT_DIR, exist_ok=True)
    with open(os.path.join(OUT_DIR, "_sheets.txt"), "w", encoding="utf-8", newline="\n") as idx:
        idx.write("\n".join(wb.sheetnames) + "\n")
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        path = os.path.join(OUT_DIR, f"{safe_name(sheet)}.tsv")
        with open(path, "w", encoding="utf-8", newline="\n") as f:
            for row in ws.iter_rows():
                line = "\t".join(cell_text(c.value) for c in row).rstrip("\t")
                f.write(line + "\n")
    print(f"Snapshot written to {OUT_DIR}/ for {len(wb.sheetnames)} sheet(s).")


if __name__ == "__main__":
    main()
